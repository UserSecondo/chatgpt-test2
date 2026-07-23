"""
===============================================================================
Proyecto : BDD_GEO_DICTIONARY3
Archivo  : excel_reader_base.py
Versión  : 5.1

Construye el modelo ProjectMetadata a partir de los archivos Excel.

Responsabilidades
-----------------
- Leer las diferentes fuentes de información.
- Construir el modelo de dominio.
- Mantener índices internos para búsquedas rápidas.

No realiza:
- Resolución de conflictos entre fuentes.
- Aplicación de reglas de negocio.
- Escritura del diccionario de datos.
===============================================================================
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Iterable, Tuple

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import Config

from metadata import (
    COLUMN_MAPS,
    INPUT_FILENAMES,
    PRIORITY,
    REQUIRED_COLUMNS,
    SHEETS,
)

from models import (
    FieldInfo,
    ProjectMetadata,
    SchemaInfo,
    TableInfo,
)


class ExcelReaderBase:
    """
    Construye el modelo ProjectMetadata leyendo todas
    las fuentes Excel del proyecto.
    """

    ###########################################################################
    # Constructor
    ###########################################################################

    def __init__(self, config: Config):

        self.config = config

        self.project = ProjectMetadata()

        #
        # Índices internos
        #

        self._schemas: Dict[str, SchemaInfo] = {}

        self._tables: Dict[Tuple[str, str], TableInfo] = {}

        self._fields: Dict[
            Tuple[str, str, str],
            FieldInfo,
        ] = {}

        #
        # RN-001: esquemas marcados para documentar
        # (columna REVISION de la hoja Usuarios).
        #

        self._documented_schemas: Dict[str, bool] = {}

    ###########################################################################
    # API pública
    ###########################################################################

    def read(self) -> ProjectMetadata:
        """
        Construye el modelo completo del proyecto.
        """

        self._load_oracle()

        self._load_inventory()

        self._load_catalog()

        self._load_esri()

        self._load_mgn()

        self._load_campos_descripciones()

        self._load_siglas_mgn()

        self._load_operaciones_dane()

        self._update_summary()

        self._validate_project()

        return self.project

    ###########################################################################
    # Apertura de archivos
    ###########################################################################

    def _open_workbook(
        self,
        filename: str,
    ) -> Workbook:
        """
        Abre un archivo Excel en modo lectura.
        """

        path = self.config.input_file(filename)

        if not path.exists():

            raise FileNotFoundError(

                f"No existe el archivo:\n{path}"

            )

        return load_workbook(

            filename=path,

            data_only=True,

            read_only=True,

        )

    ###########################################################################
    # Utilidades de lectura
    ###########################################################################

    @staticmethod
    def _worksheet_headers(
        sheet: Worksheet,
        header_row: int = 1,
    ) -> list[str]:
        """
        Obtiene los encabezados de una hoja.

        Algunas fuentes (por ejemplo el catálogo de esquemas)
        anteponen un bloque de título antes de la fila real de
        encabezados, por lo que `header_row` no siempre es 1.
        """

        headers = []

        for cell in sheet[header_row]:

            value = cell.value

            if value is None:

                headers.append("")

            else:

                headers.append(

                    str(value).strip()

                )

        return headers

    ###########################################################################

    @staticmethod
    def _normalize_text(
        value: Any,
    ) -> str:
        """
        Normaliza texto libre.

        No modifica mayúsculas/minúsculas.
        """

        if value is None:

            return ""

        return str(value).strip()

    ###########################################################################

    @staticmethod
    def _normalize_identifier(
        value: Any,
    ) -> str:
        """
        Normaliza identificadores.

        Se utiliza para:

        - esquema
        - tabla
        - campo
        """

        if value is None:

            return ""

        return str(value).strip().upper()

    ###########################################################################

    def _iter_rows(
        self,
        sheet: Worksheet,
        header_row: int = 1,
    ) -> Iterable[Dict[str, str]]:
        """
        Convierte una hoja Excel en un iterador
        de diccionarios.
        """

        headers = self._worksheet_headers(sheet, header_row)

        for row in sheet.iter_rows(

            min_row=header_row + 1,

            values_only=True,

        ):

            yield {

                headers[i]:

                self._normalize_text(value)

                for i, value in enumerate(row)

                if i < len(headers)

            }

    ###########################################################################

    @staticmethod
    def _find_header_row(
        sheet: Worksheet,
        required_columns: tuple,
        max_rows: int = 20,
    ) -> int:
        """
        Localiza la fila donde están los encabezados reales,
        buscando la primera fila que contenga todas las
        columnas obligatorias.

        Necesario porque algunas fuentes (por ejemplo el
        catálogo de esquemas) anteponen un bloque de título
        antes de los encabezados.
        """

        required_upper = {col.upper() for col in required_columns}

        for row_number in range(1, max_rows + 1):

            values = {
                str(cell.value).strip().upper()
                for cell in sheet[row_number]
                if cell.value is not None
            }

            if required_upper.issubset(values):
                return row_number

        raise ValueError(
            "No fue posible localizar la fila de encabezados "
            f"en la hoja '{sheet.title}'."
        )

    ###########################################################################
    # Validación
    ###########################################################################

    @staticmethod
    def _validate_columns(
        headers: list[str],
        required: tuple,
    ) -> None:
        """
        Verifica que existan todas las columnas
        obligatorias.
        """

        headers_upper = {

            h.upper()

            for h in headers

        }

        missing = [

            col

            for col in required

            if col.upper() not in headers_upper

        ]

        if missing:

            raise ValueError(

                "Columnas obligatorias no encontradas: "

                + ", ".join(missing)

            )

    ###########################################################################
    # Conversión al modelo lógico
    ###########################################################################

    @staticmethod
    def _logical_row(
        row: Dict[str, str],
        mapping: Dict[str, str],
    ) -> Dict[str, str]:
        """
        Convierte una fila del Excel al modelo lógico
        definido en metadata.py.
        """

        logical = {}

        for excel_name, logical_name in mapping.items():

            logical[logical_name] = row.get(
                excel_name,
                "",
            )

        return logical

    ###########################################################################
    # Lectura genérica de hojas
    ###########################################################################

    def _read_sheet(
        self,
        workbook: Workbook,
        sheet_name: str,
        mapping_name: str,
    ) -> Iterable[Dict[str, str]]:
        """
        Lee una hoja Excel y devuelve las filas
        convertidas al modelo lógico.
        """

        try:

            sheet = workbook[sheet_name]

        except KeyError as ex:

            raise ValueError(
                f"No existe la hoja '{sheet_name}'."
            ) from ex

        header_row = self._find_header_row(
            sheet,
            REQUIRED_COLUMNS[mapping_name],
        )

        headers = self._worksheet_headers(sheet, header_row)

        self._validate_columns(
            headers,
            REQUIRED_COLUMNS[mapping_name],
        )

        mapping = COLUMN_MAPS[mapping_name]

        for row in self._iter_rows(sheet, header_row):

            logical = self._logical_row(
                row,
                mapping,
            )

            #
            # Normalización de identificadores
            #

            if "schema" in logical:
                logical["schema"] = self._normalize_identifier(
                    logical["schema"]
                )

            if "table" in logical:
                logical["table"] = self._normalize_identifier(
                    logical["table"]
                )

            if "field" in logical:
                logical["field"] = self._normalize_identifier(
                    logical["field"]
                )

            yield logical

    ###########################################################################
    # Índices internos
    ###########################################################################

    def _get_schema(
        self,
        schema_name: str,
    ) -> SchemaInfo:
        """
        Obtiene un esquema existente o lo crea.
        """

        schema = self._schemas.get(schema_name)

        if schema is None:

            schema = SchemaInfo(
                schema_name=schema_name
            )

            self.project.schemas.append(schema)

            self._schemas[schema_name] = schema

        return schema

    ###########################################################################

    def _get_table(
        self,
        schema_name: str,
        table_name: str,
    ) -> TableInfo:
        """
        Obtiene una tabla existente o la crea.
        """

        key = (
            schema_name,
            table_name,
        )

        table = self._tables.get(key)

        if table is None:

            schema = self._get_schema(schema_name)

            table = TableInfo(
                schema_name=schema_name,
                table_name=table_name,
            )

            schema.tables.append(table)

            self._tables[key] = table

        return table

    ###########################################################################

    def _get_field(
        self,
        schema_name: str,
        table_name: str,
        field_name: str,
    ) -> FieldInfo:
        """
        Obtiene un campo existente o lo crea.
        """

        key = (
            schema_name,
            table_name,
            field_name,
        )

        field = self._fields.get(key)

        if field is None:

            table = self._get_table(
                schema_name,
                table_name,
            )

            field = FieldInfo(
                field_name=field_name
            )

            table.fields.append(field)

            self._fields[key] = field

        return field

    ###########################################################################
    # Asignación con reglas de prioridad entre fuentes
    ###########################################################################

    @staticmethod
    def _source_rank(source: str, priority_key: str) -> int:
        """
        Devuelve la posición de una fuente dentro del orden de
        prioridad definido en metadata.PRIORITY.

        Una fuente que no aparece en la lista se considera de
        menor prioridad que cualquiera de las listadas.
        """

        order = PRIORITY.get(priority_key, ())

        if source in order:
            return order.index(source)

        return len(order)

    def _assign_by_priority(
        self,
        obj: object,
        attribute: str,
        source_attribute: str,
        priority_key: str,
        value: str,
        source: str,
    ) -> None:
        """
        Asigna `value` a `attribute` únicamente si la fuente que lo
        provee tiene igual o mayor prioridad que la fuente que fijó
        el valor actual (o si el atributo aún no tiene fuente).

        Implementa las reglas RN-004, RN-005 y RN-006.
        """

        if not value:
            return

        current_source = getattr(obj, source_attribute, "")

        if current_source:

            current_rank = self._source_rank(current_source, priority_key)
            new_rank = self._source_rank(source, priority_key)

            if new_rank >= current_rank:
                return

        setattr(obj, attribute, value)
        setattr(obj, source_attribute, source)

    @staticmethod
    def _assign_if_empty(
        obj: object,
        attribute: str,
        value: str,
    ) -> None:
        """
        Asigna `value` únicamente si el atributo aún no tiene
        contenido. Utilizado para campos complementarios que no
        tienen una regla de prioridad explícita (por ejemplo,
        observaciones provistas tanto por ESRI como por MGN).
        """

        if not value:
            return

        if getattr(obj, attribute, ""):
            return

        setattr(obj, attribute, value)

    @staticmethod
    def _is_placeholder_description(value: str, name: str) -> bool:
        """
        RN-008: una fuente que documenta la descripción de un
        campo como el mismo nombre del campo no está aportando
        información real.
        """

        if not value or not name:
            return False

        return value.strip().upper() == name.strip().upper()

    def _is_documented(self, schema_name: str) -> bool:
        """
        Indica si un esquema está marcado para documentar
        (RN-001). Un esquema que nunca apareció en la hoja
        Usuarios se considera no documentado.
        """

        return self._documented_schemas.get(schema_name, False)

    ###########################################################################
    # Estadísticas
    ###########################################################################

    def _update_summary(self) -> None:
        """
        Actualiza el resumen de la ejecución.
        """

        summary = self.project.summary

        summary.processed_schemas = len(
            self.project.schemas
        )

        summary.processed_tables = sum(
            len(schema.tables)
            for schema in self.project.schemas
        )

        summary.processed_fields = sum(
            len(table.fields)
            for schema in self.project.schemas
            for table in schema.tables
        )
