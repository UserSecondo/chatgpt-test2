"""
===============================================================================
Proyecto : BDD_GEO_DICTIONARY3
Archivo  : excel_writer.py
Versión  : 7.0

Genera el diccionario de datos institucional a partir del modelo resuelto.

Formato de salida
------------------
- Un archivo Excel POR ESQUEMA documentado.
- Dentro de cada archivo, una hoja POR TABLA.
- Cada hoja respeta la plantilla institucional
  (Plantilla_Diccionario_Datos_DBGEODIG_AJUSTADO.xlsx):

    Esquema:
    Tabla:
    Descripción de la tabla:
    Responsable:
    Fecha de documentación:

    N° | Campo | Tipo de dato (Oracle) | Longitud/Precisión (Oracle)
       | Tipo de dato (ESRI) | Longitud/Precisión (ESRI)
       | PK | FK | Nulo | Descripción

Notas
-----
- PK y FK se dejan en blanco: ninguna fuente actual provee esa
  información (ver conversación con el usuario).
- No realiza lectura de archivos de entrada, ni reglas de negocio,
  ni resolución de conflictos entre fuentes.
===============================================================================
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from config import Config
from metadata import OUTPUT_FILENAMES, VALUES
from models import FieldInfo, ProjectMetadata, SchemaInfo, TableInfo

_TEMPLATE_SHEET_NAME = "Diccionario_Datos"

_FIRST_DATA_ROW = 8

_NO_DESCRIPTION_TEXT = "Sin Descripción"

_NO_DESCRIPTION_FILL = PatternFill(
    start_color="FFFF00",
    end_color="FFFF00",
    fill_type="solid",
)

_INVALID_SHEET_CHARS = r"[\[\]:\*\?/\\]"

_INVALID_FILENAME_CHARS = r'[\\/:*?"<>|]'


class ExcelWriter:
    """
    Genera el diccionario de datos institucional: un archivo por
    esquema, con una hoja por tabla, siguiendo la plantilla real.
    """

    def __init__(self, config: Config):
        self.config = config

    def generate(self, project: ProjectMetadata) -> list[Path]:
        """
        Genera un archivo Excel por cada esquema documentado.

        Returns
        -------
        list[Path]
            Rutas de los archivos generados.
        """

        template_path = self.config.template_file

        generated_paths: list[Path] = []

        for schema in project.schemas:

            if not schema.tables:
                continue

            output_path = self._generate_schema_workbook(
                schema,
                template_path,
            )

            generated_paths.append(output_path)

            project.summary.generated_files += 1

        missing_report_path = self._generate_missing_descriptions_report(
            project
        )

        generated_paths.append(missing_report_path)

        project.summary.generated_files += 1

        followup_report_path = self._generate_schema_followup_report(
            project
        )

        generated_paths.append(followup_report_path)

        project.summary.generated_files += 1

        project.summary.end_time = datetime.now()

        if project.summary.start_time and project.summary.end_time:
            project.summary.execution_time = (
                project.summary.end_time - project.summary.start_time
            )

        return generated_paths

    ###########################################################################
    # Reporte de seguimiento por esquema y responsable
    ###########################################################################

    def _generate_schema_followup_report(
        self,
        project: ProjectMetadata,
    ) -> Path:
        """
        Genera un archivo Excel con TODOS los esquemas de la base
        de datos (estén o no marcados para documentar), su
        responsable, y el detalle de campos sin descripción por
        esquema — pensado para solicitar aclaraciones a las áreas
        responsables.
        """

        schemas_by_name = {
            schema.schema_name: schema for schema in project.schemas
        }

        rows = []

        for schema_name, documented in sorted(
            project.all_schemas_documented.items()
        ):

            responsible = project.all_schemas_responsible.get(
                schema_name, ""
            )

            schema = schemas_by_name.get(schema_name)

            total_tables = 0
            total_fields = 0
            missing_fields = 0

            if schema is not None:

                total_tables = len(schema.tables)

                for table in schema.tables:
                    for f in table.fields:
                        total_fields += 1
                        if (
                            not f.description
                            or f.description == VALUES["unknown"]
                        ):
                            missing_fields += 1

            missing_pct = (
                round(100 * missing_fields / total_fields, 1)
                if total_fields
                else ""
            )

            rows.append(
                (
                    schema_name,
                    "SI" if documented else "NO",
                    responsible,
                    total_tables if schema is not None else "",
                    total_fields if schema is not None else "",
                    missing_fields if schema is not None else "",
                    missing_pct,
                )
            )

        missing_detail = [
            (
                schema.schema_name,
                table.table_name,
                field.field_name,
                field.data_type,
                project.all_schemas_responsible.get(
                    schema.schema_name, ""
                ),
            )
            for schema in project.schemas
            for table in schema.tables
            for field in table.fields
            if not field.description
            or field.description == VALUES["unknown"]
        ]

        workbook = Workbook()

        self._write_schema_followup_sheet(workbook.active, rows)

        self._write_missing_with_responsible_sheet(
            workbook.create_sheet("Campos Sin Descripción"),
            missing_detail,
        )

        output_path = self.config.output_file(
            OUTPUT_FILENAMES["schema_followup"]
        )

        workbook.save(output_path)
        workbook.close()

        return output_path

    @staticmethod
    def _write_schema_followup_sheet(
        sheet: Worksheet,
        rows: list[tuple],
    ) -> None:
        """
        Escribe la hoja con todos los esquemas, si están marcados
        para documentar, su responsable, y el conteo de campos
        sin descripción.
        """

        sheet.title = "Esquemas"

        headers = [
            "Esquema",
            "¿Marcado para documentar?",
            "Responsable",
            "Total tablas",
            "Total campos",
            "Campos sin descripción",
            "% sin descripción",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, name="Arial")

        for row_idx, row_values in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_values, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name="Arial")

        widths = (30, 22, 40, 14, 14, 22, 18)
        for col_idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[
                sheet.cell(row=1, column=col_idx).column_letter
            ].width = width

        sheet.freeze_panes = "A2"

    @staticmethod
    def _write_missing_with_responsible_sheet(
        sheet: Worksheet,
        missing: list[tuple[str, str, str, str, str]],
    ) -> None:
        """
        Escribe el detalle de campos sin descripción, incluyendo
        el responsable del esquema al que pertenece cada campo.
        """

        headers = ["Esquema", "Tabla", "Campo", "Tipo de dato", "Responsable"]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, name="Arial")

        for row_idx, row_values in enumerate(missing, start=2):
            for col_idx, value in enumerate(row_values, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name="Arial")

        widths = (30, 35, 30, 18, 40)
        for col_idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[
                sheet.cell(row=1, column=col_idx).column_letter
            ].width = width

        sheet.freeze_panes = "A2"

    ###########################################################################
    # Reporte de campos sin descripción
    ###########################################################################

    def _generate_missing_descriptions_report(
        self,
        project: ProjectMetadata,
    ) -> Path:
        """
        Genera un archivo Excel adicional con el detalle de los
        campos que no tienen descripción en ninguna fuente.

        Incluye una hoja de resumen por esquema y una hoja de
        detalle con cada campo (Esquema, Tabla, Campo, Tipo de dato).
        """

        missing: list[tuple[str, str, str, str]] = [
            (
                schema.schema_name,
                table.table_name,
                field.field_name,
                field.data_type,
            )
            for schema in project.schemas
            for table in schema.tables
            for field in table.fields
            if not field.description or field.description == VALUES["unknown"]
        ]

        total_fields = sum(
            len(table.fields)
            for schema in project.schemas
            for table in schema.tables
        )

        by_schema: dict[str, int] = {}
        for schema_name, *_ in missing:
            by_schema[schema_name] = by_schema.get(schema_name, 0) + 1

        workbook = Workbook()

        self._write_missing_summary_sheet(
            workbook.active,
            total_fields,
            missing,
            by_schema,
        )

        self._write_missing_detail_sheet(
            workbook.create_sheet("Detalle"),
            missing,
        )

        output_path = self.config.output_file(
            OUTPUT_FILENAMES["missing_descriptions"]
        )

        workbook.save(output_path)
        workbook.close()

        return output_path

    @staticmethod
    def _write_missing_summary_sheet(
        sheet: Worksheet,
        total_fields: int,
        missing: list[tuple[str, str, str, str]],
        by_schema: dict[str, int],
    ) -> None:
        """
        Escribe la hoja de resumen del reporte de campos sin
        descripción.
        """

        sheet.title = "Resumen"

        sheet["A1"] = "Campos sin descripción — Resumen por esquema"
        sheet["A1"].font = Font(bold=True, size=13, name="Arial")
        sheet.merge_cells("A1:B1")

        sheet["A3"] = "Total de campos documentados"
        sheet["B3"] = total_fields

        sheet["A4"] = "Total de campos sin descripción"
        sheet["B4"] = len(missing)

        sheet["A5"] = "Porcentaje sin descripción"
        sheet["B5"] = (
            round(100 * len(missing) / total_fields, 1)
            if total_fields
            else 0
        )
        sheet["C5"] = "%"

        for row in (3, 4, 5):
            sheet.cell(row=row, column=1).font = Font(name="Arial")
            sheet.cell(row=row, column=2).font = Font(
                name="Arial", bold=True
            )

        headers = ["Esquema", "Cantidad de campos sin descripción"]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=7, column=col_idx, value=header)
            cell.font = Font(bold=True, name="Arial")

        row_idx = 8
        for schema_name, count in sorted(
            by_schema.items(),
            key=lambda item: item[1],
            reverse=True,
        ):
            sheet.cell(row=row_idx, column=1, value=schema_name).font = (
                Font(name="Arial")
            )
            sheet.cell(row=row_idx, column=2, value=count).font = Font(
                name="Arial"
            )
            row_idx += 1

        sheet.column_dimensions["A"].width = 35
        sheet.column_dimensions["B"].width = 32

    @staticmethod
    def _write_missing_detail_sheet(
        sheet: Worksheet,
        missing: list[tuple[str, str, str, str]],
    ) -> None:
        """
        Escribe la hoja de detalle del reporte de campos sin
        descripción: un renglón por campo.
        """

        headers = ["Esquema", "Tabla", "Campo", "Tipo de dato"]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, name="Arial")

        for row_idx, (schema_name, table_name, field_name, data_type) in (
            enumerate(missing, start=2)
        ):
            sheet.cell(row=row_idx, column=1, value=schema_name).font = (
                Font(name="Arial")
            )
            sheet.cell(row=row_idx, column=2, value=table_name).font = (
                Font(name="Arial")
            )
            sheet.cell(row=row_idx, column=3, value=field_name).font = (
                Font(name="Arial")
            )
            sheet.cell(row=row_idx, column=4, value=data_type).font = (
                Font(name="Arial")
            )

        for col, width in zip("ABCD", (30, 35, 30, 18)):
            sheet.column_dimensions[col].width = width

        sheet.freeze_panes = "A2"

    ###########################################################################
    # Generación por esquema
    ###########################################################################

    def _generate_schema_workbook(
        self,
        schema: SchemaInfo,
        template_path: Path,
    ) -> Path:
        """
        Genera el archivo Excel de un esquema, con una hoja por tabla.
        """

        workbook = load_workbook(template_path)

        master_sheet = workbook[_TEMPLATE_SHEET_NAME]

        template_last_row = master_sheet.max_row

        used_sheet_names: set[str] = set()

        for table in schema.tables:

            sheet_name = self._unique_sheet_name(
                table.table_name,
                used_sheet_names,
            )

            sheet = workbook.copy_worksheet(master_sheet)
            sheet.title = sheet_name

            self._write_table_sheet(sheet, schema, table, template_last_row)

        # La hoja maestra y las notas de uso son insumos internos
        # de la plantilla; no forman parte del diccionario final.

        del workbook[_TEMPLATE_SHEET_NAME]

        if "Hoja2" in workbook.sheetnames:
            del workbook["Hoja2"]

        output_path = self.config.output_file(
            self._schema_filename(schema.schema_name)
        )

        workbook.save(output_path)
        workbook.close()

        return output_path

    ###########################################################################
    # Generación por tabla
    ###########################################################################

    def _write_table_sheet(
        self,
        sheet: Worksheet,
        schema: SchemaInfo,
        table: TableInfo,
        template_last_row: int,
    ) -> None:
        """
        Llena una hoja (copia de la plantilla) con la
        información de una tabla.

        El número de filas de campos queda acorde a la cantidad
        real de campos de la tabla: se eliminan las filas de
        ejemplo sobrantes cuando hay menos campos que en la
        plantilla, y se agregan filas adicionales cuando hay más.
        """

        sheet["B2"] = schema.schema_name
        sheet["B3"] = table.table_name
        sheet["B5"] = schema.responsible
        sheet["B6"] = datetime.now().strftime("%Y-%m-%d")

        if table.description and table.description != VALUES["unknown"]:
            sheet["B4"] = table.description
        else:
            sheet["B4"] = _NO_DESCRIPTION_TEXT
            sheet["B4"].fill = _NO_DESCRIPTION_FILL

        self._clear_placeholder_rows(sheet, template_last_row)

        row_index = _FIRST_DATA_ROW

        for position, field in enumerate(table.fields, start=1):
            self._write_field_row(sheet, row_index, position, field)
            row_index += 1

        last_written_row = row_index - 1

        if last_written_row < template_last_row:

            # Menos campos que filas de ejemplo en la plantilla:
            # se eliminan las filas sobrantes.

            sheet.delete_rows(
                last_written_row + 1,
                template_last_row - last_written_row,
            )

    @staticmethod
    def _write_field_row(
        sheet: Worksheet,
        row_index: int,
        position: int,
        field: FieldInfo,
    ) -> None:
        """
        Escribe una fila de campo siguiendo el orden de columnas
        de la plantilla institucional.
        """

        has_description = (
            field.description and field.description != VALUES["unknown"]
        )

        values = [
            position,
            field.field_name,
            field.oracle_type,
            field.oracle_length,
            field.esri_type,
            field.esri_length,
            "",  # PK: sin fuente disponible actualmente.
            "",  # FK: sin fuente disponible actualmente.
            field.nullable,
            field.description if has_description else _NO_DESCRIPTION_TEXT,
        ]

        for col_idx, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=col_idx, value=value)

        if not has_description:
            sheet.cell(row=row_index, column=10).fill = _NO_DESCRIPTION_FILL

    @staticmethod
    def _clear_placeholder_rows(
        sheet: Worksheet,
        template_last_row: int,
    ) -> None:
        """
        La plantilla trae filas de ejemplo numeradas bajo el
        encabezado (hasta la fila `template_last_row`). Se
        limpian antes de escribir los datos reales para no
        dejar residuos de la plantilla.
        """

        for row in range(_FIRST_DATA_ROW, template_last_row + 1):
            for col in range(1, 11):
                sheet.cell(row=row, column=col, value=None)

    ###########################################################################
    # Nombres de hoja y de archivo
    ###########################################################################

    @staticmethod
    def _unique_sheet_name(
        table_name: str,
        used_names: set[str],
    ) -> str:
        """
        Genera un nombre de hoja válido para Excel (máximo 31
        caracteres, sin caracteres inválidos) y único dentro
        del libro.
        """

        base_name = re.sub(_INVALID_SHEET_CHARS, "_", table_name)[:31]

        candidate = base_name
        suffix = 1

        while candidate in used_names:

            suffix_text = f"_{suffix}"
            candidate = base_name[: 31 - len(suffix_text)] + suffix_text
            suffix += 1

        used_names.add(candidate)

        return candidate

    @staticmethod
    def _schema_filename(schema_name: str) -> str:
        """
        Genera el nombre de archivo del diccionario de un esquema.
        """

        safe_name = re.sub(_INVALID_FILENAME_CHARS, "_", schema_name)

        return f"Diccionario_Datos_{safe_name}.xlsx"
