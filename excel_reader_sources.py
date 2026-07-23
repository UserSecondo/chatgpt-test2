"""
===============================================================================
Proyecto : BDD_GEO_DICTIONARY3
Archivo  : excel_reader_sources.py
Versión  : 5.1

Lectura de las fuentes ESRI, MGN y validación del modelo.

Responsabilidades
-----------------
- Enriquecer campos con metadata ESRI.
- Enriquecer campos con metadata MGN.
- Validar la consistencia del modelo cargado.

No realiza:
- Resolución de conflictos entre fuentes.
- Aplicación de reglas de negocio.
- Escritura del diccionario de datos.
===============================================================================
"""

from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from logger import get_logger
from metadata import SHEETS

logger = get_logger()


class ExcelReaderSources:
    """
    Implementa la lectura de ESRI, MGN y validaciones.
    """

    ###########################################################################
    # ESRI
    ###########################################################################

    def _load_esri(self) -> None:
        """
        Procesa el archivo de metadatos ESRI.
        """

        workbook = self._open_workbook("esri")

        try:
            self._read_esri_fields(workbook)
        finally:
            workbook.close()

    def _read_esri_fields(self, workbook: Workbook) -> None:
        """
        Enriquece los campos utilizando la metadata ESRI.

        Esta fuente NO crea nuevos objetos del modelo.
        Únicamente complementa la información de los
        campos ya existentes.
        """

        for row in self._read_sheet(
            workbook,
            SHEETS["esri"],
            "esri",
        ):
            schema_name = row.get("schema", "")
            table_name = row.get("table", "")
            field_name = row.get("field", "")

            if not schema_name or not table_name or not field_name:
                continue

            if not self._is_documented(schema_name):
                continue

            field = self._get_field(schema_name, table_name, field_name)

            esri_type = row.get("type", "")
            if esri_type:
                field.esri_type = esri_type

            esri_length = row.get("length", "")
            if esri_length:
                field.esri_length = str(esri_length)

            self._assign_by_priority(
                field,
                "data_type",
                "data_type_source",
                "field_description",
                esri_type,
                "esri",
            )

            self._set_if_value(
                field,
                "alias",
                row.get("alias", ""),
            )

            self._set_if_value(
                field,
                "domain",
                row.get("domain", ""),
            )

            description = row.get("description", "")

            if not self._is_placeholder_description(description, field_name):
                self._assign_by_priority(
                    field,
                    "description",
                    "description_source",
                    "field_description",
                    description,
                    "esri",
                )

            self._assign_if_empty(
                field,
                "observations",
                row.get("observations", ""),
            )

    ###########################################################################
    # MGN
    ###########################################################################

    def _load_mgn(self) -> None:
        """
        Procesa el archivo de metadatos MGN.
        """

        workbook = self._open_workbook("mgn")

        try:
            self._read_mgn(workbook)
        finally:
            workbook.close()

    def _read_mgn(self, workbook: Workbook) -> None:
        """
        Enriquece los campos utilizando la metadata MGN.

        A diferencia de Oracle o ESRI, el diccionario MGN documenta
        campos estándar únicamente por NOMBRE (por ejemplo DPTO_CCDGO,
        MPIO_CCDGO), sin indicar a qué esquema o tabla pertenecen.
        Además cada hoja contiene varios bloques de datos, cada uno
        con su propio encabezado 'CAMPOS'/'DESCRIPCIÓN' repetido.

        Por eso la enriquece de forma global: cualquier campo del
        proyecto cuyo nombre coincida recibe la descripción MGN
        correspondiente, sin importar en qué esquema o tabla esté.
        """

        mgn_descriptions: dict[str, str] = {}

        for sheet_name in SHEETS["mgn"]:

            sheet = workbook[sheet_name]

            for field_name, description in self._iter_mgn_blocks(sheet):

                if not field_name or not description:
                    continue

                if self._is_placeholder_description(description, field_name):
                    continue

                mgn_descriptions.setdefault(field_name, description)

        if not mgn_descriptions:
            return

        for schema in self.project.schemas:
            for table in schema.tables:
                for field in table.fields:

                    description = mgn_descriptions.get(field.field_name)

                    if description is None:
                        continue

                    self._assign_by_priority(
                        field,
                        "description",
                        "description_source",
                        "field_description",
                        description,
                        "mgn",
                    )

    @staticmethod
    def _iter_mgn_blocks(sheet):
        """
        Recorre una hoja MGN detectando cada bloque de datos.

        Cada hoja documenta varias feature classes seguidas, y cada
        una repite su propio encabezado 'CAMPOS' / 'DESCRIPCIÓN' en
        una posición de columna que puede variar de una hoja a otra.
        Esta función localiza cada encabezado y produce las filas de
        datos que le siguen, ignorando los encabezados repetidos.
        """

        field_col = None
        description_col = None

        for row in sheet.iter_rows(values_only=True):

            header_positions = {
                str(value).strip().upper(): idx
                for idx, value in enumerate(row)
                if value is not None
            }

            if "CAMPOS" in header_positions and "DESCRIPCIÓN" in header_positions:
                field_col = header_positions["CAMPOS"]
                description_col = header_positions["DESCRIPCIÓN"]
                continue

            if field_col is None or description_col is None:
                continue

            if field_col >= len(row) or description_col >= len(row):
                continue

            field_name = row[field_col]

            if field_name is None:
                continue

            description = row[description_col]

            yield (
                str(field_name).strip().upper(),
                str(description).strip() if description is not None else "",
            )

    ###########################################################################
    # Descripciones curadas manualmente
    ###########################################################################

    def _load_campos_descripciones(self) -> None:
        """
        Procesa el archivo de descripciones curadas manualmente
        por el equipo del proyecto, a partir del reporte de
        campos sin descripción (hoja 'Campos_Descripciones').
        """

        workbook = self._open_workbook("campos_descripciones")

        try:
            self._read_campos_descripciones(workbook)
        finally:
            workbook.close()

    def _read_campos_descripciones(self, workbook: Workbook) -> None:
        """
        Enriquece los campos con las descripciones que el equipo
        del proyecto completó manualmente. Funciona por NOMBRE de
        campo, de forma global (igual que MGN y siglas_mgn).

        Las filas marcadas explícitamente como "No Disponible" no
        aportan información real y se ignoran.
        """

        sheet = workbook[SHEETS["campos_descripciones"]]

        descriptions: dict[str, str] = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):

            if len(row) < 3:
                continue

            field_name, _, description = row[0], row[1], row[2]

            if not field_name or not description:
                continue

            field_name = str(field_name).strip().upper()
            description = str(description).strip()

            if description.upper() in ("NO DISPONIBLE", "NO DISPONIBLE."):
                continue

            if self._is_placeholder_description(description, field_name):
                continue

            descriptions.setdefault(field_name, description)

        if not descriptions:
            return

        for schema in self.project.schemas:
            for table in schema.tables:
                for field in table.fields:

                    description = descriptions.get(field.field_name)

                    if description is None:
                        continue

                    self._assign_by_priority(
                        field,
                        "description",
                        "description_source",
                        "field_description",
                        description,
                        "campos_descripciones",
                    )

    ###########################################################################
    # Diccionarios de referencia externos (info_referencia/)
    ###########################################################################

    # (ruta relativa a info_referencia/, nombre de hoja o None para todas)
    _INFO_REFERENCIA_FILES = (
        ("MGN/Diccionario_Datos_MGN_2025.xlsx", None),
        ("MGN/Diccionario_Datos_MGN_2024.xlsx", None),
        ("MGN/Diccionario_Datos_MGN_2023.xlsx", None),
        ("MGN/Diccionario_Datos_MGN_2022.xlsx", None),
        ("MGN/Diccionario_Datos_MGN_2021.xlsx", None),
        ("MGN/Diccionario_Datos_MGN_2020.xlsx", None),
        ("CNPV2018/DICCIONARIO_DATOS_CNPV2018.xlsx", "DICIONARIO"),
        (
            "CNPV2018/DICCIONARIODEDATOS_GEOCNPV_20190214_VERSION_ENTREGA.xlsx",
            "DICIONARIO DATOS",
        ),
        ("CNUE/DiccionarioDeDatos_Vistas_CNUE2021.xlsx", "ReporteAtributos"),
        ("CENU2024/Diccionario_base_estructural_UE_CENU.xlsx", "Hoja1"),
        (
            "Reg_Catastrales/MEDELLIN/Diccionario_Datos_Catastro_para_DANE_2019.xlsx",
            "GISCAT",
        ),
        (
            "MzHomologadas/DICCIONARIO_DATOS_HOMOLOGADA_2005_2019.xlsx",
            "DICCIONARIO_DATOS",
        ),
        (
            "MzHomologadas/DICCIONARIO_DATOS_HOMOLOGADA_2020_2021.xlsx",
            "DICCIONARIO_DATOS",
        ),
        (
            "MzHomologadas/DICCIONARIO_DATOS_HOMOLOGADA_2021_2022.xlsx",
            "DICCIONARIO_DATOS",
        ),
        (
            "Reg_Catastrales/MEDELLIN/MEDELLIN_DICCIONARIO_DATOS_2015.xlsx",
            "Hoja1",
        ),
        (
            "Reg_Catastrales/IGAC/igac - DICCIONARIO DE DATOS.xlsx",
            "diccionario datos",
        ),
        (
            "Reg_Catastrales/Gestores_Catastrales/20220930_DICCIONARIO_MOCE2021.xlsx",
            "DICCIONARIO",
        ),
        ("MMRA/Diccionario_Datos_MMRA_2017.xlsx", "MMRA_2017"),
        ("MMRA/Diccionario_Datos_MMRA_2018.xlsx", "MMRA_2017"),
    )

    _INFO_REFERENCIA_FIELD_ALIASES = {
        "CAMPO",
        "CAMPOS",
        "VARIABLE",
        "ATRIBUTO",
        "NOMBRE CAMPO",
    }

    _INFO_REFERENCIA_DESCRIPTION_ALIASES = {
        "DESCRIPCION",
        "DESCRIPCIÓN",
        "DESCRIPCIÓN BREVE",
        "DESCRIPCION BREVE",
        "CORRESPONDE A",
    }

    def _load_info_referencia(self) -> None:
        """
        Procesa los diccionarios de referencia externos que el
        equipo del proyecto reunió en la carpeta info_referencia/
        (diccionarios de otras operaciones y entidades: MGN
        2020-2025, CNPV2018, CNUE2021, CENU2024, manzanas
        homologadas, catastros municipales, MMRA).

        Esta carpeta es opcional y no se distribuye en el
        repositorio (pesa varios cientos de MB): si no está
        presente, este paso simplemente se omite.
        """

        base_dir = self.config.BASE_DIR / "info_referencia"

        if not base_dir.exists():
            logger.warning(
                "Carpeta info_referencia/ no encontrada; se omite "
                "el enriquecimiento con diccionarios de referencia "
                "externos."
            )
            return

        descriptions: dict[str, str] = {}

        for relative_path, sheet_name in self._INFO_REFERENCIA_FILES:

            file_path = base_dir / relative_path

            if not file_path.exists():
                logger.warning(
                    "No se encontró '%s' dentro de info_referencia/; "
                    "se omite este archivo.",
                    relative_path,
                )
                continue

            workbook = load_workbook(
                file_path,
                read_only=True,
                data_only=True,
            )

            try:
                sheet_names = (
                    [sheet_name] if sheet_name else workbook.sheetnames
                )

                for name in sheet_names:

                    sheet = workbook[name]

                    for field_name, description in (
                        self._iter_generic_field_blocks(
                            sheet,
                            self._INFO_REFERENCIA_FIELD_ALIASES,
                            self._INFO_REFERENCIA_DESCRIPTION_ALIASES,
                        )
                    ):
                        if not field_name or not description:
                            continue

                        if self._is_placeholder_description(
                            description,
                            field_name,
                        ):
                            continue

                        descriptions.setdefault(field_name, description)
            finally:
                workbook.close()

        if not descriptions:
            return

        for schema in self.project.schemas:
            for table in schema.tables:
                for field in table.fields:

                    description = descriptions.get(field.field_name)

                    if description is None:
                        continue

                    self._assign_by_priority(
                        field,
                        "description",
                        "description_source",
                        "field_description",
                        description,
                        "info_referencia",
                    )

    ###########################################################################
    # Diccionario de siglas MGN
    ###########################################################################

    def _load_siglas_mgn(self) -> None:
        """
        Procesa el diccionario de siglas MGN (glosario adicional
        de nombres de campo estándar, más amplio que las hojas
        MGN originales).
        """

        workbook = self._open_workbook("siglas_mgn")

        try:
            self._read_siglas_mgn(workbook)
        finally:
            workbook.close()

    def _read_siglas_mgn(self, workbook: Workbook) -> None:
        """
        Enriquece los campos utilizando el diccionario de siglas
        MGN. Al igual que las hojas MGN, funciona por NOMBRE de
        campo, de forma global (sin esquema ni tabla).
        """

        sheet = workbook[SHEETS["siglas_mgn"]]

        descriptions: dict[str, str] = {}

        for sigla, nombre in sheet.iter_rows(min_row=2, values_only=True):

            if not sigla or not nombre:
                continue

            sigla = str(sigla).strip().upper()
            nombre = str(nombre).strip()

            if self._is_placeholder_description(nombre, sigla):
                continue

            descriptions.setdefault(sigla, nombre)

        if not descriptions:
            return

        for schema in self.project.schemas:
            for table in schema.tables:
                for field in table.fields:

                    description = descriptions.get(field.field_name)

                    if description is None:
                        continue

                    self._assign_by_priority(
                        field,
                        "description",
                        "description_source",
                        "field_description",
                        description,
                        "siglas_mgn",
                    )

    ###########################################################################
    # Catálogo de operaciones DANE (heurística de siglas)
    ###########################################################################

    def _load_operaciones_dane(self) -> None:
        """
        Procesa el catálogo de operaciones estadísticas del DANE
        para inferir la descripción de esquemas y tablas cuyo
        nombre contenga la sigla de una operación reconocida
        (RN-009).

        Esta es una inferencia, no una fuente oficial: solo se
        aplica cuando ninguna otra fuente aportó una descripción.
        """

        workbook = self._open_workbook("operaciones_dane")

        try:
            operations = self._read_operaciones_dane(workbook)
        finally:
            workbook.close()

        if not operations:
            return

        for schema in self.project.schemas:

            if not schema.description:
                inferred = self._infer_from_operations(
                    schema.schema_name,
                    operations,
                )

                if inferred:
                    schema.description = inferred
                    schema.description_source = "operaciones_dane"

            for table in schema.tables:

                if table.description:
                    continue

                inferred = self._infer_from_operations(
                    table.table_name,
                    operations,
                )

                if inferred:
                    table.description = inferred
                    table.description_source = "operaciones_dane"

    @staticmethod
    def _read_operaciones_dane(workbook: Workbook) -> dict[str, str]:
        """
        Lee el catálogo de operaciones DANE y devuelve un
        diccionario SIGLA -> NOMBRE de la operación.
        """

        sheet = workbook[SHEETS["operaciones_dane"]]

        operations: dict[str, str] = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):

            if len(row) < 3:
                continue

            _, nombre, sigla = row[0], row[1], row[2]

            if not sigla or not nombre:
                continue

            operations[str(sigla).strip().upper()] = str(nombre).strip()

        return operations

    @staticmethod
    def _infer_from_operations(
        name: str,
        operations: dict[str, str],
    ) -> str:
        """
        Intenta inferir una descripción a partir de los tokens
        (separados por '_') del nombre de un esquema o tabla,
        buscando coincidencia exacta con una sigla de operación
        conocida.

        Para evitar inferencias ambiguas, solo se acepta cuando
        hay exactamente una sigla reconocida entre los tokens.
        """

        tokens = name.upper().split("_")

        matches = [token for token in tokens if token in operations]

        if len(matches) != 1:
            return ""

        acronym = matches[0]
        full_name = operations[acronym]

        period_tokens = [
            token
            for token in tokens
            if token != acronym and token.isdigit()
        ]

        if period_tokens:
            return f"{full_name} ({'-'.join(period_tokens)})"

        return full_name

    ###########################################################################
    # Validación del modelo
    ###########################################################################

    def _validate_project(self) -> None:
        """
        Ejecuta todas las validaciones del modelo cargado.
        """

        self._validate_schemas()
        self._validate_tables()
        self._validate_fields()

    def _validate_schemas(self) -> None:
        """
        Valida la información de los esquemas.
        """

        for schema in self.project.schemas:
            if not schema.tables:
                self.project.warnings.append(
                    f"El esquema '{schema.schema_name}' no contiene tablas."
                )

    def _validate_tables(self) -> None:
        """
        Valida la información de las tablas.
        """

        for schema in self.project.schemas:
            for table in schema.tables:
                if not table.fields:
                    self.project.warnings.append(
                        f"La tabla '{table.table_name}' no contiene campos."
                    )

    def _validate_fields(self) -> None:
        """
        Valida los campos del modelo.
        """

        for schema in self.project.schemas:
            for table in schema.tables:
                for field in table.fields:
                    if not field.data_type:
                        self.project.warnings.append(
                            f"{schema.schema_name}.{table.table_name}.{field.field_name} no tiene tipo de dato."
                        )

                    if not field.description:
                        self.project.warnings.append(
                            f"{schema.schema_name}.{table.table_name}.{field.field_name} no tiene descripción."
                        )
