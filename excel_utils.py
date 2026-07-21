"""
===============================================================================
Proyecto : BDD_GEO_DICTIONARY3
Archivo  : excel_utils.py
Autor    : Nelson David Martínez
Versión  : 3.0

Descripción
-----------
Biblioteca de utilidades para el manejo de archivos Excel.

Responsabilidades
-----------------
- Abrir libros Excel.
- Obtener hojas.
- Detectar encabezados.
- Convertir filas en diccionarios.
- Normalizar valores.
- Validar columnas.
- Funciones reutilizables para lectura y escritura.

Este módulo NO contiene lógica de negocio.
===============================================================================
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Iterator, List

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelUtils:

    """
    Biblioteca estática para trabajar con Excel.
    """

    # -------------------------------------------------------------------------
    # Libros
    # -------------------------------------------------------------------------

    @staticmethod
    def load_workbook(
        file_path: Path,
        read_only: bool = True,
        data_only: bool = True
    ) -> Workbook:
        """
        Abre un archivo Excel.

        Parameters
        ----------
        file_path : Path

        Returns
        -------
        Workbook
        """

        if not file_path.exists():

            raise FileNotFoundError(file_path)

        return load_workbook(
            filename=file_path,
            read_only=read_only,
            data_only=data_only
        )

    # -------------------------------------------------------------------------
    # Hojas
    # -------------------------------------------------------------------------

    @staticmethod
    def get_sheet(
        workbook: Workbook,
        sheet_name: str
    ) -> Worksheet:

        if sheet_name not in workbook.sheetnames:

            raise ValueError(
                f"La hoja '{sheet_name}' no existe."
            )

        return workbook[sheet_name]

    # -------------------------------------------------------------------------
    # Encabezados
    # -------------------------------------------------------------------------

    @staticmethod
    def read_headers(
        sheet: Worksheet,
        header_row: int
    ) -> Dict[str, int]:

        headers = {}

        for index, cell in enumerate(sheet[header_row]):

            if cell.value is None:
                continue

            value = str(cell.value).strip()

            headers[value] = index

        return headers

    # -------------------------------------------------------------------------

    @staticmethod
    def find_header_row(
        sheet: Worksheet,
        required_columns: List[str],
        max_rows: int = 20
    ) -> int:
        """
        Busca automáticamente la fila donde están
        los encabezados.
        """

        for row in range(1, max_rows + 1):

            headers = ExcelUtils.read_headers(
                sheet,
                row
            )

            if all(
                column in headers
                for column in required_columns
            ):

                return row

        raise ValueError(
            "No fue posible localizar los encabezados."
        )

    # -------------------------------------------------------------------------
    # Filas
    # -------------------------------------------------------------------------

    @staticmethod
    def iter_rows(
        sheet: Worksheet,
        header_row: int
    ) -> Iterator[Dict[str, Any]]:

        headers = ExcelUtils.read_headers(
            sheet,
            header_row
        )

        for row in sheet.iter_rows(
            min_row=header_row + 1,
            values_only=True
        ):

            if ExcelUtils.is_empty_row(row):

                continue

            record = {}

            for column, index in headers.items():

                value = None

                if index < len(row):

                    value = ExcelUtils.normalize(
                        row[index]
                    )

                record[column] = value

            yield record

    # -------------------------------------------------------------------------
    # Utilidades
    # -------------------------------------------------------------------------

    @staticmethod
    def normalize(
        value: Any
    ) -> Any:

        if value is None:

            return None

        if isinstance(value, str):

            value = value.strip()

            if value == "":

                return None

        return value

    # -------------------------------------------------------------------------

    @staticmethod
    def safe_get(
        row: Dict[str, Any],
        column: str,
        default=None
    ):

        return row.get(column, default)

    # -------------------------------------------------------------------------

    @staticmethod
    def is_empty_row(
        row
    ) -> bool:

        return all(
            ExcelUtils.normalize(v) is None
            for v in row
        )

    # -------------------------------------------------------------------------

    @staticmethod
    def worksheet_dimension(
        sheet: Worksheet
    ) -> tuple[int, int]:

        return (
            sheet.max_row,
            sheet.max_column
        )

    # -------------------------------------------------------------------------

    @staticmethod
    def validate_required_columns(
        headers: Dict[str, int],
        required_columns: List[str]
    ):

        missing = [

            column

            for column in required_columns

            if column not in headers

        ]

        if missing:

            raise ValueError(

                "Faltan columnas obligatorias: "

                + ", ".join(missing)

            )