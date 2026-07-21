"""
===============================================================================
Proyecto : BDD_GEO_DICTIONARY3
Archivo  : excel_reader.py
Versión  : 5.1

Construye el modelo ProjectMetadata a partir de los archivos Excel.

Responsabilidades
-----------------
- Leer las diferentes fuentes de información.
- Construir el modelo de dominio.
- Mantener índices internos para búsquedas rápidas.

No realiza:
- Resolución de conflictos entre fuentes.
- Aplicación de reglas de negocio.
- Escritura del diccionario de datos.
===============================================================================
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Iterable, Tuple

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import Config

from metadata import (
    COLUMN_MAPS,
    INPUT_FILENAMES,
    REQUIRED_COLUMNS,
    SHEETS,
)

from models import (
    FieldInfo,
    ProjectMetadata,
    SchemaInfo,
    TableInfo,
)


class ExcelReader:
    """
    Construye el modelo ProjectMetadata leyendo todas
    las fuentes Excel del proyecto.
    """

    ###########################################################################
    # Constructor
    ###########################################################################

    def __init__(self, config: Config):

        self.config = config

        self.project = ProjectMetadata()

        #
        # Índices internos
        #

        self._schemas: Dict[str, SchemaInfo] = {}

        self._tables: Dict[Tuple[str, str], TableInfo] = {}

        self._fields: Dict[
            Tuple[str, str, str],
            FieldInfo
        ] = {}

    ###########################################################################
    # API pública
    ###########################################################################

    def read(self) -> ProjectMetadata:
        """
        Construye el modelo completo del proyecto.
        """

        self._load_oracle()

        self._load_inventory()

        self._load_catalog()

        self._load_esri()

        self._load_mgn()

        self._update_summary()

        self._validate_project ()

        return self.project

    ###########################################################################
    # Apertura de archivos
    ###########################################################################

    def _open_workbook(
        self,
        filename: str
    ) -> Workbook:
        """
        Abre un archivo Excel en modo lectura.
        """

        path = self.config.input_file(filename)

        if not path.exists():

            raise FileNotFoundError(

                f"No existe el archivo:\n{path}"

            )

        return load_workbook(

            filename=path,

            data_only=True,

            read_only=True

        )

    ###########################################################################
    # Utilidades de lectura
    ###########################################################################

    @staticmethod
    def _worksheet_headers(
        sheet: Worksheet
    ) -> list[str]:
        """
        Obtiene los encabezados de una hoja.
        """

        headers = []

        for cell in sheet[1]:

            value = cell.value

            if value is None:

                headers.append("")

            else:

                headers.append(

                    str(value).strip()

                )

        return headers

    ###########################################################################

    @staticmethod
    def _normalize_text(
        value: Any
    ) -> str:
        """
        Normaliza texto libre.

        No modifica mayúsculas/minúsculas.
        """

        if value is None:

            return ""

        return str(value).strip()

    ###########################################################################

    @staticmethod
    def _normalize_identifier(
        value: Any
    ) -> str:
        """
        Normaliza identificadores.

        Se utiliza para:

        - esquema
        - tabla
        - campo
        """

        if value is None:

            return ""

        return str(value).strip().upper()

    ###########################################################################

    def _iter_rows(
        self,
        sheet: Worksheet
    ) -> Iterable[Dict[str, str]]:
        """
        Convierte una hoja Excel en un iterador
        de diccionarios.
        """

        headers = self._worksheet_headers(sheet)

        for row in sheet.iter_rows(

            min_row=2,

            values_only=True

        ):

            yield {

                headers[i]:

                self._normalize_text(value)

                for i, value in enumerate(row)

            }

    ###########################################################################
    # Validación
    ###########################################################################

    @staticmethod
    def _validate_columns(
        headers: list[str],
        required: tuple
    ) -> None:
        """
        Verifica que existan todas las columnas
        obligatorias.
        """

        headers_upper = {

            h.upper()

            for h in headers

        }

        missing = [

            col

            for col in required

            if col.upper() not in headers_upper

        ]

        if missing:

            raise ValueError(

                "Columnas obligatorias no encontradas: "

                + ", ".join(missing)

            )
    ###########################################################################
    # Conversión al modelo lógico
    ###########################################################################

    @staticmethod
    def _logical_row(
        row: Dict[str, str],
        mapping: Dict[str, str]
    ) -> Dict[str, str]:
        """
        Convierte una fila del Excel al modelo lógico
        definido en metadata.py.
        """

        logical = {}

        for excel_name, logical_name in mapping.items():

            logical[logical_name] = row.get(
                excel_name,
                ""
            )

        return logical

    ###########################################################################
    # Lectura genérica de hojas
    ###########################################################################

    def _read_sheet(
        self,
        workbook: Workbook,
        sheet_name: str,
        mapping_name: str,
    ) -> Iterable[Dict[str, str]]:
        """
        Lee una hoja Excel y devuelve las filas
        convertidas al modelo lógico.
        """

        try:

            sheet = workbook[sheet_name]

        except KeyError as ex:

            raise ValueError(
                f"No existe la hoja '{sheet_name}'."
            ) from ex

        headers = self._worksheet_headers(sheet)

        self._validate_columns(
            headers,
            REQUIRED_COLUMNS[mapping_name]
        )

        mapping = COLUMN_MAPS[mapping_name]

        for row in self._iter_rows(sheet):

            logical = self._logical_row(
                row,
                mapping
            )

            #
            # Normalización de identificadores
            #

            if "schema" in logical:
                logical["schema"] = self._normalize_identifier(
                    logical["schema"]
                )

            if "table" in logical:
                logical["table"] = self._normalize_identifier(
                    logical["table"]
                )

            if "field" in logical:
                logical["field"] = self._normalize_identifier(
                    logical["field"]
                )

            yield logical

    ###########################################################################
    # Índices internos
    ###########################################################################

    def _get_schema(
        self,
        schema_name: str
    ) -> SchemaInfo:
        """
        Obtiene un esquema existente o lo crea.
        """

        schema = self._schemas.get(schema_name)

        if schema is None:

            schema = SchemaInfo(
                schema_name=schema_name
            )

            self.project.schemas.append(schema)

            self._schemas[schema_name] = schema

        return schema

    ###########################################################################

    def _get_table(
        self,
        schema_name: str,
        table_name: str
    ) -> TableInfo:
        """
        Obtiene una tabla existente o la crea.
        """

        key = (
            schema_name,
            table_name
        )

        table = self._tables.get(key)

        if table is None:

            schema = self._get_schema(schema_name)

            table = TableInfo(
                schema_name=schema_name,
                table_name=table_name
            )

            schema.tables.append(table)

            self._tables[key] = table

        return table

    ###########################################################################

    def _get_field(
        self,
        schema_name: str,
        table_name: str,
        field_name: str
    ) -> FieldInfo:
        """
        Obtiene un campo existente o lo crea.
        """

        key = (
            schema_name,
            table_name,
            field_name
        )

        field = self._fields.get(key)

        if field is None:

            table = self._get_table(
                schema_name,
                table_name
            )

            field = FieldInfo(
                field_name=field_name
            )

            table.fields.append(field)

            self._fields[key] = field

        return field

    ###########################################################################
    # Estadísticas
    ###########################################################################

    def _update_summary(self) -> None:
        """
        Actualiza el resumen de la ejecución.
        """

        summary = self.project.summary

        summary.processed_schemas = len(
            self.project.schemas
        )

        summary.processed_tables = sum(
            len(schema.tables)
            for schema in self.project.schemas
        )

        summary.processed_fields = sum(
            len(table.fields)
            for schema in self.project.schemas
            for table in schema.tables
        )

    ###########################################################################
    # Métodos de carga
    ###########################################################################

     def _load_catalog(self) -> None:
        """Implementado en la Parte 3."""
        pass

    def _load_esri(self) -> None:
        """Implementado en la Parte 3."""
        pass

    def _load_mgn(self) -> None:
        """Implementado en la Parte 3."""
        pass
    ###########################################################################
    # Oracle
    ###########################################################################

    def _load_oracle(self) -> None:
        """
        Procesa el archivo de metadatos Oracle.
        """

        workbook = self._open_workbook(

            INPUT_FILENAMES["oracle"]

        )

        try:

            self._read_users(workbook)

            self._read_tables(workbook)

            self._read_oracle_fields(workbook)

        finally:

            workbook.close()

    ###########################################################################

    def _read_users(
        self,
        workbook: Workbook
    ) -> None:
        """
        Lee la hoja de usuarios Oracle.

        Crea los esquemas e incorpora el responsable
        cuando esté disponible.
        """

        for row in self._read_sheet(

            workbook,

            SHEETS["users"],

            "users"

        ):

            schema_name = row.get(

                "schema",

                ""

            )

            if not schema_name:

                continue

            schema = self._get_schema(

                schema_name

            )

            responsible = row.get(

                "responsible",

                ""

            )

            if responsible:

                schema.responsible = responsible

                schema.responsible_source = "oracle"
    ###########################################################################
    # Tablas Oracle
    ###########################################################################

    def _read_tables(
        self,
        workbook: Workbook
    ) -> None:
        """
        Lee la hoja de tablas Oracle.
        """

        for row in self._read_sheet(

            workbook,

            SHEETS["tables"],

            "tables"

        ):

            schema_name = row.get(

                "schema",

                ""

            )

            table_name = row.get(

                "table",

                ""

            )

            if not schema_name or not table_name:

                continue

            table = self._get_table(

                schema_name,

                table_name

            )

            description = row.get(

                "description",

                ""

            )

            if description:

                table.description = description

                table.description_source = "oracle"

    ###########################################################################
    # Campos Oracle
    ###########################################################################

    def _read_oracle_fields(
        self,
        workbook: Workbook
    ) -> None:
        """
        Lee los campos Oracle.
        """

        for row in self._read_sheet(

            workbook,

            SHEETS["oracle_fields"],

            "oracle_fields"

        ):

            schema_name = row.get(

                "schema",

                ""

            )

            table_name = row.get(

                "table",

                ""

            )

            field_name = row.get(

                "field",

                ""

            )

            if (

                not schema_name

                or not table_name

                or not field_name

            ):

                continue

            field = self._get_field(

                schema_name,

                table_name,

                field_name

            )

            #
            # Tipo de dato
            #

            value = row.get(

                "type",

                ""

            )

            if value:

                field.data_type = value

                field.data_type_source = "oracle"

            #
            # Longitud
            #

            value = row.get(

                "length",

                ""

            )

            if value:

                field.length = value

            #
            # Nullable
            #

            value = row.get(

                "nullable",

                ""

            )

            if value:

                field.nullable = value

            #
            # Descripción
            #

            value = row.get(

                "description",

                ""

            )

            if value:

                field.description = value

                field.description_source = "oracle"
    ###########################################################################
    # Utilidad de asignación
    ###########################################################################

    @staticmethod
    def _set_if_value(
        obj: object,
        attribute: str,
        value: str,
        source_attribute: str | None = None,
        source: str | None = None,
    ) -> None:
        """
        Asigna un valor únicamente cuando éste no es vacío.

        Si se indica un atributo de origen, también registra
        la fuente del dato.
        """

        if value is None:
            return

        value = value.strip()

        if not value:
            return

        setattr(
            obj,
            attribute,
            value
        )

        if source_attribute and source:

            setattr(
                obj,
                source_attribute,
                source
            )

    ###########################################################################
    # Inventario Oracle
    ###########################################################################

    def _load_inventory(self) -> None:
        """
        Procesa el inventario Oracle.
        """

        workbook = self._open_workbook(

            INPUT_FILENAMES["inventory"]

        )

        try:

            self._read_inventory(workbook)

        finally:

            workbook.close()

    ###########################################################################

    def _read_inventory(
        self,
        workbook: Workbook
    ) -> None:
        """
        Lee la información del inventario Oracle.
        """

        for row in self._read_sheet(

            workbook,

            SHEETS["inventory"],

            "inventory"

        ):

            schema_name = row.get(

                "schema",

                ""

            )

            table_name = row.get(

                "table",

                ""

            )

            if not schema_name or not table_name:

                continue

            table = self._get_table(

                schema_name,

                table_name

            )

            #
            # Descripción
            #

            self._set_if_value(

                table,

                "description",

                row.get(

                    "description",

                    ""

                ),

                "description_source",

                "inventory"

            )

    ###########################################################################
    # Catálogo
    ###########################################################################

    def _load_catalog(self) -> None:
        """
        Procesa el catálogo de esquemas.
        """

        workbook = self._open_workbook(

            INPUT_FILENAMES["catalog"]

        )

        try:

            self._read_catalog(workbook)

        finally:

            workbook.close()

    ###########################################################################

    def _read_catalog(
        self,
        workbook: Workbook
    ) -> None:
        """
        Lee el catálogo de esquemas.
        """

        for row in self._read_sheet(

            workbook,

            SHEETS["catalog"],

            "catalog"

        ):

            schema_name = row.get(

                "schema",

                ""

            )

            if not schema_name:

                continue

            schema = self._get_schema(

                schema_name

            )

            #
            # Responsable
            #

            self._set_if_value(

                schema,

                "responsible",

                row.get(

                    "responsible",

                    ""

                ),

                "responsible_source",

                "catalog"

            )

            #
            # Descripción
            #

            self._set_if_value(

                schema,

                "description",

                row.get(

                    "description",

                    ""

                ),

                "description_source",

                "catalog"

            )
    ###########################################################################
    # ESRI
    ###########################################################################

    def _load_esri(self) -> None:
        """
        Procesa el archivo de metadatos ESRI.
        """

        workbook = self._open_workbook(

            INPUT_FILENAMES["esri"]

        )

        try:

            self._read_esri_fields(

                workbook

            )

        finally:

            workbook.close()

    ###########################################################################

    def _read_esri_fields(
        self,
        workbook: Workbook
    ) -> None:
        """
        Enriquece los campos utilizando la metadata ESRI.

        Esta fuente NO crea nuevos objetos del modelo.
        Únicamente complementa la información de los
        campos ya existentes.
        """

        for row in self._read_sheet(

            workbook,

            SHEETS["esri"],

            "esri"

        ):

            schema_name = row.get(

                "schema",

                ""

            )

            table_name = row.get(

                "table",

                ""

            )

            field_name = row.get(

                "field",

                ""

            )

            if (

                not schema_name

                or not table_name

                or not field_name

            ):

                continue

            field = self._get_field(

                schema_name,

                table_name,

                field_name

            )

            ###################################################################
            # Alias
            ###################################################################

            self._set_if_value(

                field,

                "alias",

                row.get(

                    "alias",

                    ""

                )

            )

            ###################################################################
            # Dominio
            ###################################################################

            self._set_if_value(

                field,

                "domain",

                row.get(

                    "domain",

                    ""

                )

            )

            ###################################################################
            # Descripción
            ###################################################################

            self._set_if_value(

                field,

                "description",

                row.get(

                    "description",

                    ""

                ),

                "description_source",

                "esri"

            )

            ###################################################################
            # Observaciones
            ###################################################################

            self._set_if_value(

                field,

                "observations",

                row.get(

                    "observations",

                    ""

                )

            )
    ###########################################################################
    # MGN
    ###########################################################################

    def _load_mgn(self) -> None:
        """
        Procesa el archivo de metadatos MGN.
        """

        workbook = self._open_workbook(

            INPUT_FILENAMES["mgn"]

        )

        try:

            self._read_mgn(

                workbook

            )

        finally:

            workbook.close()

    ###########################################################################

    def _read_mgn(
        self,
        workbook: Workbook
    ) -> None:
        """
        Enriquece los campos utilizando la metadata MGN.

        Esta fuente complementa la información existente,
        sin aplicar reglas de prioridad.
        """

        for row in self._read_sheet(

            workbook,

            SHEETS["mgn"],

            "mgn"

        ):

            schema_name = row.get(

                "schema",

                ""

            )

            table_name = row.get(

                "table",

                ""

            )

            field_name = row.get(

                "field",

                ""

            )

            if (

                not schema_name

                or not table_name

                or not field_name

            ):

                continue

            field = self._get_field(

                schema_name,

                table_name,

                field_name

            )

            ###################################################################
            # Descripción
            ###################################################################

            self._set_if_value(

                field,

                "description",

                row.get(

                    "description",

                    ""

                ),

                "description_source",

                "mgn"

            )

            ###################################################################
            # Observaciones
            ###################################################################

            self._set_if_value(

                field,

                "observations",

                row.get(

                    "observations",

                    ""

                )

            )
###########################################################################
# Validación del modelo
###########################################################################

    def _validate_project(self) -> None:
        """
    Ejecuta todas las validaciones del modelo cargado.
        """

        self._validate_schemas()

        self._validate_tables()

        self._validate_fields()

###########################################################################

def _validate_schemas(self) -> None:
    """
    Valida la información de los esquemas.
    """

    for schema in self.project.schemas.values():

        if not schema.tables:

            self.project.warnings.append(

                f"El esquema '{schema.schema_name}' no contiene tablas."

            )

###########################################################################

def _validate_tables(self) -> None:
    """
    Valida la información de las tablas.
    """

    for schema in self.project.schemas.values():

        for table in schema.tables.values():

            if not table.fields:

                self.project.warnings.append(

                    f"La tabla '{table.table_name}' no contiene campos."

                )

###########################################################################

def _validate_fields(self) -> None:
    """
    Valida los campos del modelo.
    """

    for schema in self.project.schemas.values():

        for table in schema.tables.values():

            for field in table.fields.values():

                if not field.data_type:

                    self.project.warnings.append(

                        f"{schema.schema_name}.{table.table_name}.{field.field_name} "
                        "no tiene tipo de dato."

                    )

                if not field.description:

                    self.project.warnings.append(

                        f"{schema.schema_name}.{table.table_name}.{field.field_name} "
                        "no tiene descripción."

                    )
